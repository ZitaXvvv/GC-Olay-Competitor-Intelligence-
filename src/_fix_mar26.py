"""
修复 CI_List_Ada Mar'26.xlsx：提取真实数据行，重新保存干净文件
"""
import openpyxl
from pathlib import Path

DOCS = Path(r"C:\Users\xie.x.3\Documents\Olay CI")
fname = "CI_List_Ada Mar'26.xlsx"
p = DOCS / fname

wb_in  = openpyxl.load_workbook(p, read_only=True)
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

HEADERS = ['upload time','Name','English / Benefit','Notification Time','#',
           'Registration Time','Ingredient','link','化妆品产品标签样稿','mini POC']

total = 0
for sh in wb_in.sheetnames:
    ws_in  = wb_in[sh]
    ws_out = wb_out.create_sheet(sh)
    for col, h in enumerate(HEADERS, 1):
        ws_out.cell(1, col, h)
    row_out = 2
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        if not row[1]:   # no name
            continue
        for col, val in enumerate(row, 1):
            ws_out.cell(row_out, col, val)
        row_out += 1
        total += 1
    print(f"  [{sh}] kept {row_out - 2} rows")

wb_in.close()
wb_out.save(p)
wb_out.close()
print(f"\n[DONE] Cleaned {fname}: {total} real rows saved")
