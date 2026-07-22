"""
模块3：从 Excel 读取 PDF URL，下载 Artwork 文件到本地

对应原 UiBot 流程块2.task 的 PDF 下载逻辑：
  - 先清空品牌本地文件夹（与 UiBot 行为一致）
  - 对时间窗口内的每条记录，读 H 列 (link) 的 URL
  - URL 含 hzpba.nmpa.gov.cn → 普通化妆品备案  → 保存为 {Name}.pdf
  - URL 含 nmpa.gov.cn/datasearch → 特殊化妆品注册 → 保存为 特化--{Name}.pdf

下载策略：
  hzpba：先 HTTP 直下；失败则用 Playwright 从页面提取 PDF URL
  nmpa datasearch：从 ?url= 参数提取真实 PDF 地址，再 HTTP 下载
"""

import logging
import re
import shutil
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import requests
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from config import (
    BRANDS,
    COL_NAME,
    COL_PDF_URL,
    COL_UPLOAD_DATE,
    DOWNLOAD_BASE,
    EXCEL_PATH,
    TIME_PERIOD_DAYS,
)

log = logging.getLogger(__name__)

HZPBA_HOST   = "hzpba.nmpa.gov.cn"
NMPA_HOST    = "nmpa.gov.cn"

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/pdf,*/*",
}


def sanitize(name: str) -> str:
    """把产品名转为合法文件名"""
    return re.sub(r'[\\/:*?"<>|\r\n]', "_", name.strip())[:120]


def _is_pdf_response(resp) -> bool:
    ct = resp.headers.get("Content-Type", "")
    return "text/html" not in ct and len(resp.content) > 1024


def download_file(url: str, save_path: Path, referer: str = "",
                  cookies: dict | None = None) -> bool:
    """HTTP 直接下载，返回是否成功（SSL 验证关闭，兼容公司网络）"""
    headers = dict(REQUEST_HEADERS)
    if referer:
        headers["Referer"] = referer
    try:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

        resp = requests.get(url, headers=headers, cookies=cookies or {},
                            timeout=90, stream=True, verify=False)
        resp.raise_for_status()

        if not _is_pdf_response(resp):
            log.warning(f"  响应不是PDF（Content-Type={resp.headers.get('Content-Type','')}）: {url[:80]}")
            return False

        save_path.parent.mkdir(parents=True, exist_ok=True)
        with open(save_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                f.write(chunk)
        size_kb = save_path.stat().st_size // 1024
        log.info(f"  ✅ 下载完成: {save_path.name} ({size_kb} KB)")
        return True

    except requests.HTTPError as e:
        log.error(f"  HTTP {e.response.status_code}: {url[:80]}")
        return False
    except Exception as e:
        log.error(f"  下载出错: {e}")
        return False


def resolve_nmpa_pdf_url(preview_url: str) -> Optional[str]:
    """
    从 nmpa.gov.cn/datasearch/preview-pdf.html?url=... 提取真实 PDF 地址。
    示例: .../preview-pdf.html?url=https://www.nmpa.gov.cn/.../file.pdf
    """
    if "url=" in preview_url:
        actual = preview_url.split("url=", 1)[1].split("&")[0]
        return actual
    return preview_url


def download_hzpba(url: str, save_path: Path) -> bool:
    """
    下载 hzpba 普通化妆品备案 PDF。
    gsxxFilePreview 会直接触发浏览器文件下载（Content-Disposition: attachment），
    不是普通网页，因此：
      - 直接 HTTP GET → 400（服务器要求会话 Cookie）
      - Playwright page.goto → "Download is starting"（需要 accept_downloads=True）
    用 Playwright expect_download() 捕获下载并保存。
    """
    from playwright.sync_api import sync_playwright

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=["--disable-blink-features=AutomationControlled"],
            )
            ctx = browser.new_context(
                accept_downloads=True,
                locale="zh-CN",
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )
            page = ctx.new_page()

            # 先访问主页建立 Session Cookie
            try:
                page.goto("https://hzpba.nmpa.gov.cn/", timeout=20000)
                page.wait_for_timeout(1000)
            except Exception:
                pass

            # 用 expect_download 捕获触发的文件下载
            with page.expect_download(timeout=60000) as dl_info:
                page.goto(url, timeout=60000,
                          wait_until="commit")   # commit = 收到响应头即可

            dl = dl_info.value
            save_path.parent.mkdir(parents=True, exist_ok=True)
            dl.save_as(str(save_path))
            browser.close()

            size_kb = save_path.stat().st_size // 1024
            log.info(f"  ✅ 下载完成: {save_path.name} ({size_kb} KB)")
            return True

    except Exception as e:
        log.error(f"  Playwright 下载出错: {e}")

    return False


def _hzpba_download_with_page(page, url: str, save_path: Path) -> bool:
    """用已有 Playwright page 对象下载单个 hzpba PDF。
    hzpba 的 gsxxFilePreview 端点直接返回 Content-Disposition:attachment，
    page.goto() 会抛出 "Download is starting"；
    在 with expect_download() 内 catch 掉该异常，退出 with 后 dl_info.value 即可拿到下载对象。
    """
    try:
        with page.expect_download(timeout=60000) as dl_info:
            try:
                page.goto(url, timeout=60000, wait_until="commit")
            except Exception:
                pass  # "Download is starting" 是预期行为，下载事件已触发

        dl = dl_info.value
        save_path.parent.mkdir(parents=True, exist_ok=True)
        dl.save_as(str(save_path))
        size_kb = save_path.stat().st_size // 1024
        log.info(f"  ✅ 下载完成: {save_path.name} ({size_kb} KB)")
        return True
    except Exception as e:
        log.error(f"  下载失败: {e}")
        return False


def run():
    """
    遍历 Excel 所有品牌 sheet，下载时间窗口内有 PDF URL 的文件。
    每次运行前清空品牌本地文件夹（与原 UiBot 行为一致）。
    hzpba 下载使用单个 Playwright 实例（避免事件循环冲突）。
    """
    today = date.today()

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    brand_tasks: dict[str, list[dict]] = {b: [] for b in BRANDS}

    for sheet_name in wb.sheetnames:
        if sheet_name not in BRANDS:
            continue
        ws = wb[sheet_name]
        for row_idx in range(2, ws.max_row + 1):
            upload_val = ws.cell(row=row_idx, column=COL_UPLOAD_DATE).value
            name_val   = ws.cell(row=row_idx, column=COL_NAME).value
            pdf_val    = ws.cell(row=row_idx, column=COL_PDF_URL).value

            if not name_val or not pdf_val or str(pdf_val).strip() in ("", "NA"):
                continue

            # 时间窗口：兼容 datetime 对象和字符串两种格式
            try:
                if hasattr(upload_val, "date"):
                    upload_date = upload_val.date()
                elif isinstance(upload_val, date):
                    upload_date = upload_val
                else:
                    upload_date = None
                    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
                        try:
                            upload_date = datetime.strptime(str(upload_val).strip(), fmt).date()
                            break
                        except ValueError:
                            pass
                if upload_date is None:
                    continue
                if (today - upload_date).days >= TIME_PERIOD_DAYS:
                    continue
            except Exception:
                continue

            url_str   = str(pdf_val).strip()
            safe_name = sanitize(str(name_val))
            filename  = f"{safe_name}.pdf" if HZPBA_HOST in url_str else f"特化--{safe_name}.pdf"

            brand_tasks[sheet_name].append({
                "name":      str(name_val),
                "url":       url_str,
                "save_path": DOWNLOAD_BASE / sheet_name / filename,
            })

    wb.close()

    total = sum(len(v) for v in brand_tasks.values())
    if total == 0:
        log.info("无需下载的文件")
        return

    # 清空品牌文件夹（与 UiBot 一致）
    for brand, tasks in brand_tasks.items():
        if not tasks:
            continue
        brand_dir = DOWNLOAD_BASE / brand
        if brand_dir.exists():
            shutil.rmtree(brand_dir)
            log.info(f"已清空文件夹: {brand_dir}")
        brand_dir.mkdir(parents=True, exist_ok=True)

    log.info(f"待下载: {total} 个文件")

    # 分类：hzpba 用 Playwright；nmpa datasearch 用 requests
    hzpba_tasks = [(b, t) for b, ts in brand_tasks.items() for t in ts if HZPBA_HOST in t["url"]]
    nmpa_tasks  = [(b, t) for b, ts in brand_tasks.items() for t in ts if HZPBA_HOST not in t["url"]]

    success = failed = 0

    # ── NMPA datasearch（特殊化妆品）：直接 HTTP 下载 ──
    for brand, task in nmpa_tasks:
        url = task["url"]
        save_path = task["save_path"]
        log.info(f"[{brand}] {task['name']}")
        real_url = resolve_nmpa_pdf_url(url)
        if real_url and real_url != url:
            log.info(f"  解析PDF: {real_url[:90]}")
        ok = download_file(real_url or url, save_path, referer="https://www.nmpa.gov.cn/")
        if ok: success += 1
        else:
            failed += 1
            log.warning(f"  ❌ 失败: {task['name']}")

    # ── hzpba（普通化妆品）：单个 Playwright 实例，逐个 expect_download ──
    if hzpba_tasks:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=["--disable-blink-features=AutomationControlled"],
            )
            ctx = browser.new_context(
                accept_downloads=True,
                locale="zh-CN",
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
            )
            page = ctx.new_page()

            # 先访问主页建立 Session Cookie（一次即可）
            try:
                page.goto("https://hzpba.nmpa.gov.cn/", timeout=20000)
                page.wait_for_timeout(1000)
                log.info("hzpba Session 已建立")
            except Exception as e:
                log.warning(f"hzpba 主页访问失败（继续尝试下载）: {e}")

            for brand, task in hzpba_tasks:
                log.info(f"[{brand}] {task['name']}")
                ok = _hzpba_download_with_page(page, task["url"], task["save_path"])
                if ok: success += 1
                else:
                    failed += 1
                    log.warning(f"  ❌ 失败: {task['name']}")

            browser.close()

    log.info(f"\n✅ 模块3 完成 | 成功: {success}  失败: {failed}")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    run()
