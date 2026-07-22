"""
模块5：统计各品牌产品数量，发送汇总邮件
- 优先用 Outlook COM（本机安装 Outlook 时更可靠，与原 UiBot 一致）
- 若 Outlook 不可用，则改用 Microsoft Graph API 发送
"""

import logging
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
import requests

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent / "extend" / "python"))

import msal
from config import (
    BRANDS,
    COL_PDF_URL,
    COL_REG_NUM,
    COL_UPLOAD_DATE,
    EMAIL_FROM,
    EMAIL_SUBJECT,
    EMAIL_TO,
    EXCEL_PATH,
    SP_CLIENT_ID,
    SP_CLIENT_SECRET,
    SP_DETAIL_LINK,
    SP_TENANT_ID,
    TIME_PERIOD_DAYS,
)

log = logging.getLogger(__name__)

# 真实备案/注册号正则模式（防止将"特定宣称"误判为特殊化妆品）
_RE_SPECIAL = re.compile(
    r"(国妆特字|国妆特进字|卫妆特字|进口化妆品注册证号)",
    re.IGNORECASE,
)
_RE_NORMAL = re.compile(
    r"(妆网备字|国妆备字|国妆备进字|省妆备字|\w{1,4}G妆网备字)",
    re.IGNORECASE,
)


def _find_reg_in_row(row: tuple) -> tuple[str, str]:
    """
    在行的所有列里搜索真实备案/注册号（用正则匹配，避免误判功效文字）。
    返回 (reg_num, reg_type)；找不到则返回 ("", "")。
    """
    for cell in row:
        s = str(cell or "").strip()
        if not s or len(s) < 6:
            continue
        if _RE_SPECIAL.search(s):
            return s, "特殊注册"
        if _RE_NORMAL.search(s):
            return s, "普通备案"
    return "", ""


# ─────────────────────────────────────────────
# 统计
# ─────────────────────────────────────────────

def count_new_products() -> dict:
    """
    返回 {brand_en: [normal_count, special_count]}。
    - 时间过滤：A列 upload_time 在窗口内（兼容 datetime 对象和字符串）
    - 分类：用正则在整行搜索真实备案号（避免功效文字误判）
    """
    today = date.today()
    counts = {}

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    for brand_en in BRANDS:
        if brand_en not in wb.sheetnames:
            counts[brand_en] = [0, 0]
            continue

        ws = wb[brand_en]
        normal = special = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            upload_raw = row[COL_UPLOAD_DATE - 1] if len(row) >= COL_UPLOAD_DATE else None

            if not upload_raw:
                continue

            # 时间过滤
            try:
                if hasattr(upload_raw, "date"):
                    upload_date = upload_raw.date()
                elif isinstance(upload_raw, date):
                    upload_date = upload_raw
                else:
                    upload_date = None
                    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
                        try:
                            upload_date = datetime.strptime(str(upload_raw).strip(), fmt).date()
                            break
                        except ValueError:
                            pass
                if upload_date is None or (today - upload_date).days >= TIME_PERIOD_DAYS:
                    continue
            except Exception:
                continue

            # 搜索整行找真实备案/注册号（正则匹配，防误判）
            _, reg_type = _find_reg_in_row(row)
            if reg_type == "特殊注册":
                special += 1
            elif reg_type == "普通备案":
                normal += 1

        counts[brand_en] = [normal, special]

    wb.close()
    return counts


# ─────────────────────────────────────────────
# 邮件 HTML
# ─────────────────────────────────────────────

def build_html(counts: dict) -> str:
    today_str = date.today().strftime("%Y-%m-%d")
    rows_html = ""
    for brand_en in BRANDS:
        normal, special = counts.get(brand_en, [0, 0])
        rows_html += f"""
    <tr>
      <td>{brand_en}</td>
      <td style="text-align:center;">{special}</td>
      <td style="text-align:center;">{normal}</td>
    </tr>"""

    return f"""<!DOCTYPE html>
<html>
<head>
  <style>
    body  {{ font-family: Arial, sans-serif; font-size: 14px; }}
    table {{ border-collapse: collapse; width: 60%; margin-top: 10px; }}
    td, th {{ border: 1px solid #dddddd; text-align: left; padding: 8px; }}
    th {{ background-color: #d9e1f2; }}
    a  {{ color: #1F5C99; }}
  </style>
</head>
<body>
  <h4>Dear GC OLAY SI Team,</h4>
  <p>Please refer to the table below for new products completed notification/registration
  from key competitors in the past <b>{TIME_PERIOD_DAYS} days</b> (as of {today_str}).</p>

  <table>
    <tr>
      <th>Brand</th>
      <th>Special Cosmetics (注册)</th>
      <th>Normal Cosmetics (备案)</th>
    </tr>
    {rows_html}
  </table>

  <br>
  <p>For full details (ingredients, efficacy claims, artwork PDF), please visit:</p>
  <p>
    📂 <b>Product Detail on SharePoint:</b>
    <a href="{SP_DETAIL_LINK}">点击查看 →</a>
  </p>

  <br>
  <p style="color: #888; font-size: 12px;">
    Sent by automated Python CI bot (replacing UiBot) · {today_str}
  </p>
</body>
</html>"""


# ─────────────────────────────────────────────
# 发送方式1：Outlook COM（推荐，无需额外权限）
# ─────────────────────────────────────────────

def send_via_outlook(html_body: str) -> bool:
    """使用本机 Outlook COM 对象发送邮件（要求 Outlook 已安装并登录）"""
    try:
        import win32com.client
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)   # 0 = olMailItem
        mail.Subject  = EMAIL_SUBJECT
        mail.HTMLBody = html_body
        mail.To       = "; ".join(EMAIL_TO)
        # 附带 Excel 文件
        if Path(EXCEL_PATH).exists():
            mail.Attachments.Add(EXCEL_PATH)
        mail.Send()
        log.info(f"✅ 邮件已通过 Outlook 发送 → {EMAIL_TO}")
        return True
    except ImportError:
        log.warning("win32com 未安装，跳过 Outlook 方式")
        return False
    except Exception as exc:
        log.error(f"Outlook 发送失败: {exc}")
        return False


# ─────────────────────────────────────────────
# 发送方式2：Microsoft Graph API（备用）
# ─────────────────────────────────────────────

def _get_graph_token() -> Optional[str]:
    authority = f"https://login.microsoftonline.com/{SP_TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        SP_CLIENT_ID,
        client_credential=SP_CLIENT_SECRET,
        authority=authority,
    )
    result = app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" in result:
        return result["access_token"]
    log.error(f"Graph Token 获取失败: {result.get('error_description')}")
    return None


def send_via_graph(html_body: str) -> bool:
    """
    通过 Microsoft Graph API 发送邮件。
    【前提】：Azure AD 应用需要具有 Mail.Send 权限（应用权限，非委托权限）。
    如果尚未配置，请使用 Outlook COM 方式（send_via_outlook）。
    """
    token = _get_graph_token()
    if not token:
        return False

    to_list = [{"emailAddress": {"address": addr}} for addr in EMAIL_TO]
    payload = {
        "message": {
            "subject": EMAIL_SUBJECT,
            "body": {"contentType": "HTML", "content": html_body},
            "from": {"emailAddress": {"address": EMAIL_FROM}},
            "toRecipients": to_list,
        },
        "saveToSentItems": True,
    }

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    try:
        resp = requests.post(
            f"https://graph.microsoft.com/v1.0/users/{EMAIL_FROM}/sendMail",
            json=payload,
            headers=headers,
            timeout=30,
        )
        if resp.status_code == 202:
            log.info(f"✅ 邮件已通过 Graph API 发送 → {EMAIL_TO}")
            return True
        log.error(f"Graph 发送失败: {resp.status_code} {resp.text[:200]}")
        return False
    except Exception as exc:
        log.error(f"Graph 发送异常: {exc}")
        return False


# ─────────────────────────────────────────────
# 主入口
# ─────────────────────────────────────────────

def run():
    log.info("统计各品牌产品数量...")
    counts = count_new_products()

    for brand, (normal, special) in counts.items():
        log.info(f"  {brand}: 普通={normal}, 特殊={special}")

    html = build_html(counts)

    # 优先 Outlook COM，失败则尝试 Graph API
    if not send_via_outlook(html):
        log.info("切换为 Graph API 发送邮件...")
        if not send_via_graph(html):
            raise RuntimeError("邮件发送失败（Outlook COM 和 Graph API 均不可用）")

    log.info("✅ 模块5 完成")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    run()
