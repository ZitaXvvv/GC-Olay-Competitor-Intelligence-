"""
模块2：补全普通化妆品的省级备案 PDF URL
- 从 Excel 读取没有 PDF URL 的普通化妆品记录
- 在 hzpba.nmpa.gov.cn 搜索，提取 attachmentId
- 构造 PDF 预览 URL 并写回 Excel
"""

import logging
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
from playwright.sync_api import Page, sync_playwright

sys.path.insert(0, str(Path(__file__).parent))
from config import (
    COL_DATE,
    COL_NAME,
    COL_PDF_URL,
    COL_REG_NUM,
    COL_UPLOAD_DATE,
    EXCEL_PATH,
    HZPBA_IMPORT_URL,
    HZPBA_PDF_BASE,
    HZPBA_SEARCH_URL,
    TIME_PERIOD_DAYS,
)

log = logging.getLogger(__name__)


def search_hzpba(page: Page, reg_num: str, product_name: str, is_imported: bool = False) -> Optional[str]:
    """
    在 hzpba.nmpa.gov.cn 搜索普通化妆品备案，返回 PDF URL。
    根据 is_imported 选择国产备案页或进口备案页。
    URL 格式: https://hzpba.nmpa.gov.cn/.../gsxxFilePreview?attachmentId=XXXXXXXXXX
    """
    search_url = HZPBA_IMPORT_URL if is_imported else HZPBA_SEARCH_URL
    _origin = "进口" if is_imported else "国产"
    log.info(f"  hzpba 查询 ({_origin}): {reg_num} / {product_name}")
    try:
        page.goto(search_url, timeout=30000)
        page.wait_for_timeout(3000)

        # 输入产品名称
        search_box = page.locator(
            "input[placeholder*='产品名称'], "
            "input[placeholder*='请输入'], "
            ".el-input__inner"
        ).first
        search_box.fill(product_name)

        # 点查询
        page.locator(
            "button:has-text('查询'), button:has-text('搜索'), "
            ".el-button--primary"
        ).first.click()
        page.wait_for_timeout(3000)

        # 检查是否有结果
        no_data = page.locator(
            "text=暂无数据, text=未查到, .el-table__empty-block, .no-data"
        ).first
        if no_data.count() > 0:
            log.info(f"  hzpba 无结果: {product_name}")
            return None

        # 找到与备案号匹配的行（如果备案号已知）
        rows = page.locator("tr.el-table__row, .el-table__body tr").all()
        target_row = None
        for row in rows:
            row_text = row.inner_text()
            if reg_num and reg_num in row_text:
                target_row = row
                break
        if target_row is None and rows:
            target_row = rows[0]   # fallback：选第一条

        if not target_row:
            return None

        # 点击该行进入详情
        target_row.click()
        page.wait_for_timeout(3000)

        # ── 方式1：页面中存在 attachmentId 的直接链接 ──
        for link in page.locator("a[href*='attachmentId'], a[href*='gsxxFilePreview']").all():
            href = link.get_attribute("href") or ""
            if href:
                if href.startswith("/"):
                    href = "https://hzpba.nmpa.gov.cn" + href
                log.info(f"  找到备案PDF: {href[:80]}")
                return href

        # ── 方式2：从 embed/iframe src 提取 ──
        for el in page.locator("embed[src], iframe[src]").all():
            src = el.get_attribute("src") or ""
            if "attachmentId" in src or "gsxxFilePreview" in src:
                return src

        # ── 方式3：从页面 HTML 中正则提取 attachmentId ──
        html = page.content()
        match = re.search(r"attachmentId[=\s:\"']+(\d{15,20})", html)
        if match:
            attachment_id = match.group(1)
            url = f"{HZPBA_PDF_BASE}?attachmentId={attachment_id}"
            log.info(f"  从页面提取到 attachmentId: {attachment_id}")
            return url

        # ── 方式4：从页面 URL 提取（有时跳转后 URL 含 attachmentId）──
        current = page.url
        m = re.search(r"attachmentId=(\d+)", current)
        if m:
            return f"{HZPBA_PDF_BASE}?attachmentId={m.group(1)}"

        log.warning(f"  详情页未找到PDF: {product_name}")
        return None

    except Exception as exc:
        log.error(f"hzpba 查询出错（{reg_num} / {product_name}）: {exc}", exc_info=True)
        return None


def run():
    """从 Excel 读取需要补全 URL 的普通化妆品，查询 hzpba 后回写 Excel"""
    today = date.today()
    cutoff = today - timedelta(days=TIME_PERIOD_DAYS)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    tasks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx in range(2, ws.max_row + 1):
            upload_val  = ws.cell(row=row_idx, column=COL_UPLOAD_DATE).value  # A upload time
            name_val    = ws.cell(row=row_idx, column=COL_NAME).value          # B Name
            reg_val     = ws.cell(row=row_idx, column=COL_REG_NUM).value       # F Registration
            pdf_val     = ws.cell(row=row_idx, column=COL_PDF_URL).value       # H link

            if not name_val:
                continue
            # 只处理还没有 PDF URL 且备案号含“备”的行
            if pdf_val:
                continue
            reg_str = str(reg_val).strip() if reg_val else ""
            if "备" not in reg_str:   # 特殊化妆品（含“特”）由模块1处理，跳过
                continue

            # 检查时间窗口（按 upload time 列，A 列）
            try:
                prod_date = datetime.strptime(str(upload_val), "%m/%d/%Y").date()
                if prod_date < cutoff:
                    continue
            except Exception:
                continue

            is_imported = "进" in reg_str
            tasks.append({
                "sheet":       sheet_name,
                "row":         row_idx,
                "name":        str(name_val).strip(),
                "reg_num":     reg_str,
                "is_imported": is_imported,
            })

    if not tasks:
        log.info("无需补全的普通化妆品 URL")
        return

    log.info(f"待查询 hzpba: {len(tasks)} 条")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--lang=zh-CN"],
        )
        ctx = browser.new_context(locale="zh-CN")
        page = ctx.new_page()

        for item in tasks:
            _orig = "进口" if item["is_imported"] else "国产"
            log.info(f"查询 hzpba [{_orig}]: 【{item['sheet']}】{item['name']} ({item['reg_num']})")
            url = search_hzpba(page, item["reg_num"], item["name"], item["is_imported"])

            ws = wb[item["sheet"]]
            if url:
                ws.cell(row=item["row"], column=COL_PDF_URL, value=url)
                log.info(f"  ✅ URL 已更新: {url[:70]}")
            else:
                ws.cell(row=item["row"], column=COL_PDF_URL, value="NA")
                log.warning(f"  ⚠️ 未找到URL，标记为 NA")

        browser.close()

    wb.save(EXCEL_PATH)
    log.info("✅ 模块2 完成，Excel 已更新")


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    run()
